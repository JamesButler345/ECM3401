import networkx as nx
import sys
import openpyxl
import statistics
import spacy
from unidecode import unidecode
import os
from scipy.stats import kruskal
import json
import requests
import re

def valence_scores(directory, semantic_analysis_path, tag):

    if tag == 'true':
        nlp = spacy.load('it_core_news_sm')
    else:
        nlp = spacy.load('en_core_web_sm', disable=['parser', 'ner'])

    w_wb = openpyxl.Workbook()
    w_sheet = w_wb.active

    w_sheet.cell(row = 1, column = 1).value = "Word"
    w_sheet.cell(row = 1, column = 2).value = "Sematic"
    w_sheet.cell(row = 1, column = 3).value = "Scores (1-5)"
    w_sheet.cell(row = 1, column = 4).value = "Word Score Median"
    w_sheet.cell(row = 1, column = 5).value = "File Score Median"
    w_sheet.cell(row = 1, column = 6).value = "Kruskal p value"

    # Dictionary to hold nodes and valence scores of network
    word_dict = {}

    # List to store valence scores of entire network
    file_scores = list()


    # American to English Translater
    url = "https://raw.githubusercontent.com/hyperreality/American-British-English-Translator/master/data/american_spellings.json"
    american_to_british_dict = requests.get(url).json()

    # Filter to remove Special Char Words
    string_check= re.compile('[@_!#$%^&*()<>?/\|}{~:]+')

    for filename in os.listdir(directory):
        fileName = os.path.join(directory, filename)
        # checking if it is a file
        if os.path.isfile(fileName):

            wb_obj = openpyxl.load_workbook(fileName)
            # Select Sheet
            sheet = wb_obj[wb_obj.sheetnames[0]]

            # Count for No. lines
            count = 0

            # Iterate through rows of the current .xlsx file
            for row in sheet.iter_rows():
                count += 1
                if count == 20:
                    break

                if row[0].value == None:
                    continue

                currentNode = ''

                for cell in row:

                    if cell.value == None:
                        break

                    # Remove floats
                    cell.value = str(cell.value)
                    if '.' in cell.value:
                        word = cell.value
                        word = word.split('.', 1)
                        cell.value = word[0]

                    # Removing Special Chars
                    if string_check.search(cell.value) != None:
                        continue

                    # Apply valence score to currentNode value in word_dict
                    if cell.value.isnumeric() == True:
                        cell.value = int(cell.value)
                        word_dict[currentNode].append(cell.value)
                        file_scores.append(cell.value)
                        continue
                
                    # If node in have a space, replace with underscore
                    if ' ' in cell.value:
                        word = cell.value
                        word = word.strip()
                        cell.value = word
                        if ' ' in word:
                            word = word.replace(' ', '_')
                            cell.value = word

                    # Remove capitals and special chars
                    currentNode = cell.value
                    currentNode = currentNode.lower()

                    # American English to British English Translater
                    if tag != 'true':
                        if currentNode in american_to_british_dict:
                            currentNode = re.sub(f'(?<![a-zA-Z]){currentNode}(?![a-z-Z])', american_to_british_dict[currentNode], currentNode)

                    # Lemmatise
                    doc = nlp(currentNode)
                    currentNode = doc[0].lemma_

                    # Remove Italian accented characters
                    if tag == 'true':
                        currentNode = unidecode(currentNode)

                    # If word not already in word_dict add word and set it's value to a list
                    if currentNode not in word_dict:
                        word_dict[currentNode] = list()

    # Dictionary to hold valences scores of nodes
    valence_dict = {}

    # Initialise row and column count for writing to excel
    w_row = 2
    w_col = 1

    for item in word_dict:

        # Make copy of network scores for manipulation
        file_scores_save = file_scores.copy()

        # Remove current node's scores from file_scores_save for accurate avg calculation
        for score in word_dict[item]:
            file_scores_save.remove(score)

        # Handle edge case where no valence score was given by surveyee
        if len(word_dict[item]) == 0:
            valence_dict[item] = 'neutral'
            continue

        # Current nodes avg score
        item_median = statistics.median(word_dict[item])
    
        # Networks avg score (with current node's scores removed)  
        file_median = statistics.median(file_scores_save)

        # Run kruskal stats test on current node comparing to entire network
        stat, p = kruskal(word_dict[item], file_scores_save)

        # Confidence level for stats test
        conf_level = 0.1

        # Apply valence category
        if len(word_dict[item]) < 4:
            if item_median < 3:
                valence_dict[item] = 'negative'
            elif item_median > 3:
                valence_dict[item] = 'positive'
            else:
                valence_dict[item] = 'neutral'
        elif p > conf_level:
            valence_dict[item] = 'neutral'
        elif p < conf_level and item_median == file_median:
            valence_dict[item] = 'neutral'
        elif p < conf_level and item_median < file_median:
            valence_dict[item] = 'negative'
        elif p < conf_level and item_median > file_median:
            valence_dict[item] = 'positive'

        # Writing Semantic Anaylsis to excel file
        w_sheet.cell(row = w_row, column = w_col).value = item
        w_col += 1
        w_sheet.cell(row = w_row, column = w_col).value = str(valence_dict[item])
        w_col += 1
        w_sheet.cell(row = w_row, column = w_col).value = str(word_dict[item])
        w_col += 1
        w_sheet.cell(row = w_row, column = w_col).value = item_median
        w_col += 1
        w_sheet.cell(row = w_row, column = w_col).value = file_median
        w_col += 1
        w_sheet.cell(row = w_row, column = w_col).value = p
        w_col -= 5
        w_row += 1

        w_wb.save(valence_analysis_path)


    return valence_dict, word_dict, nlp


def create_network_closeness_file(G):

    # Initialise new workbook
    w_wb = openpyxl.Workbook()
    w_sheet = w_wb.active

    w_row = 1
    w_col = 1

    # Create edge list of network
    G = nx.read_edgelist("../py/node_pairs.txt")

    # Collate closeness scores of network
    closeness_cen = nx.closeness_centrality(G)

    # Sort closeness scores in decending order
    sortedlist = sorted(closeness_cen.items(), key=lambda item: item[1], reverse=True)

    # write headings for network_closeness_scores.xlsx
    w_sheet.cell(row = w_row, column = w_col).value = "Nodes"
    w_col += 1
    w_sheet.cell(row = w_row, column = w_col).value = "Closeness score"
    w_row += 1
    w_col -= 1

    for item in sortedlist:
    
        # Write centrality scores to network closeness scores.xlsx
        w_sheet.cell(row = w_row, column = w_col).value = item[0]
        w_col += 1
        w_sheet.cell(row = w_row, column = w_col).value = item[1]
        w_row += 1
        w_col -= 1

    w_wb.save("../py/downloads/network_closeness_scores.xlsx")


def write_dataset(valence_scores, G, root, nlp, dataset_path):
    nodesP = []
    for n in G.neighbors(root):
         nodesP.append(n)
    
    closeness_cen = nx.closeness_centrality(G)
    
    file = open(dataset_path, "w")

    file.write("[" +"\n")

    # LINKS BETWEEN NEIGHBOURS OF ROOT NODE
    for n in nodesP:
        first = True
        file.write('{"name":"flare.' +valence_scores[n] +'.' +str(n) +'","size":10,"semantic":"' +valence_scores[n] +'","closeness":"'+ str(round(closeness_cen[n],4)) +'","imports":[' )

        for m in G.neighbors(n):
            for node in nodesP:
                if m == node:
                    if first:
                        first = False
                        file.write('"flare.' +valence_scores[m] +'.' +str(m) +'"' )
                    else:
                        file.write(',"flare.' +valence_scores[m] +'.' +str(m) +'"')

        file.write(']},' + "\n")

    # LINKS TO ROOT NODE
    first = True
    doc4 = nlp(root)
    file.write('{"name":"flare.' +valence_scores[doc4[0].lemma_] +'.' +str(doc4[0].lemma_) +'","size":"root","semantic":"' +valence_scores[doc4[0].lemma_] +'","closeness":"'+ str(round(closeness_cen[root],4)) +'","imports":[')
    for n in nodesP:
        if first:
            first = False
            file.write('"flare.' +valence_scores[n] +'.' +str(n) +'"')
        else:
            file.write(',"flare.' +valence_scores[n] +'.' +str(n) +'"')

    file.write("]}]")

if __name__ == "__main__":

    directory = "../wp-content/uploads/wp_dndcf7_uploads/wpcf7-files"
    valence_analysis_path = "../py/downloads/node_valence_analysis.xlsx"

    tag = sys.argv[2]

    valence_scores, word_scores, nlp = valence_scores(directory, valence_analysis_path, tag)


    G = nx.read_edgelist("../py/node_pairs.txt")

    create_network_closeness_file(G)

    root = sys.argv[1]

    dataset_path = "../py/dataset/dataset"


    write_dataset(valence_scores, G, root, nlp, dataset_path)

    print(json.dumps(word_scores))

