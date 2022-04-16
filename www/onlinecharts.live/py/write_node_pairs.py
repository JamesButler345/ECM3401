import sys
import openpyxl
import os
import spacy
from unidecode import unidecode
import requests
import re

tag = sys.argv[1]

print("inside write nodes")

# Initialise Lemmatizer

if tag == 'true':
    nlp = spacy.load('it_core_news_sm')
else:
    nlp = spacy.load('en_core_web_sm', disable=['parser', 'ner'])

print("insde write 2")
# Initialise new workbook
w_wb = openpyxl.Workbook()
w_sheet = w_wb.active

# Node Pairs Location
file = open("../py/node_pairs.txt", "w")

# User Uploaded File location
directory = '../wp-content/uploads/wp_dndcf7_uploads/wpcf7-files'
print("inside write after directory select")
w_row = 1
w_col = 1

# American to English Translater
url = "https://raw.githubusercontent.com/hyperreality/American-British-English-Translator/master/data/american_spellings.json"
american_to_british_dict = requests.get(url).json()

# Remove Special Char Words
string_check= re.compile('[@_!#$%^&*()<>?/\|}{~:]+')

for filename in os.listdir(directory):
    fileName = os.path.join(directory, filename)
    # checking if it is a file
    if os.path.isfile(fileName):

        wb_obj = openpyxl.load_workbook(fileName)
        sheet = wb_obj[wb_obj.sheetnames[0]]

        count = 0

        for row in sheet.iter_rows():
            count += 1

            if count == 20:
                break

            if row[0].value == None:
                continue

            homeNode = ''

            for cell in row:

                if cell.value == None:
                    break

                cell.value = str(cell.value)
                if '.' in cell.value:
                    word = cell.value
                    word = word.split('.', 1)
                    cell.value = word[0]

                # Skipping valence scores in xlsx uploads
                if cell.value.isnumeric() == True:
                    continue

                # Removing Special Chars
                if string_check.search(cell.value) != None:
                    continue

                # Replacing spaces with underscores
                if ' ' in cell.value:
                    word = cell.value
                    word = word.strip()
                    cell.value = word
                    if ' ' in word:
                        word = word.replace(' ', '_')
                        cell.value = word

                # Store primary node
                if cell == row[0]:
                    homeNode = cell.value
                    continue

                # Remove capitals
                homeNode = homeNode.lower()
                word = cell.value
                word = word.lower()

                # American English to British English Translater
                if tag != 'true':
                    if homeNode in american_to_british_dict:
                        homeNode = re.sub(f'(?<![a-zA-Z]){homeNode}(?![a-z-Z])', american_to_british_dict[homeNode], homeNode)
                    if word in american_to_british_dict:
                        word = re.sub(f'(?<![a-zA-Z]){word}(?![a-z-Z])', american_to_british_dict[word], word)
                
                # Lemmatise primary node
                doc = nlp(homeNode)
                homeNode = doc[0].lemma_

                # Remove italian characters
                if tag == 'true':
                    homeNode = unidecode(homeNode)

                # Write homeNode to node_pairs.txt and lemmatized_node_pairs.xlsx
                w_sheet.cell(row = w_row, column = w_col).value = homeNode
                w_col += 1
                file.write(homeNode + '  ')
               
                # Lemmatize secondary nodes
                doc1 = nlp(word)
                word = doc1[0].lemma_

                # Remove italian characters
                if tag == 'true':
                    word = unidecode(word)

                # Write secondary nodes to node_pairs.txt and lemmatized_node_pairs.xlsx
                w_sheet.cell(row = w_row, column = w_col).value = word
                w_row += 1
                w_col -= 1
                file.write(word + "\n")



w_wb.save("../py/downloads/lemmatized_node_pairs.xlsx")

print("end of write nodes")
